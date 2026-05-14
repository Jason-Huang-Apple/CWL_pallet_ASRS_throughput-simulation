'''
1. 定义priority
2. 定义对应小车容积和波次
3. 定义波次

'''
import sys

from openpyxl import load_workbook
from collections import defaultdict
import math




def remove_duplicate(li):
    li = [i for i in li if str(i).strip()!='']
    unique_list = list(dict.fromkeys(li))
    print(unique_list)
    # print(unique_list)
    return unique_list

def count_work_center_order(res_work_sheet):
    '''

    for each work center, find each work center(priority)'s order qty
    :param res_work_sheet:
    :return: dict, key = work center, value = qty
    '''
    # 基于 res_work_sheet 的结果，按 Work Center 分组并统计去重后的 Order 数量
    work_center_orders_count = {}

    for col_name, values in res_work_sheet.items():
        # 假设 'Work Center' 和 'Order' 列存在于数据中
        if col_name == 'Work center':
            work_centers = values
            # 找到 'Order' 列的数据
            orders = res_work_sheet.get('Order', [])

            # 确保两个列表长度一致
            min_length = min(len(work_centers), len(orders))

            # 创建一个临时字典来存储每个 Work Center 对应的唯一 Orders
            temp_dict = {}
            for i in range(min_length):
                wc = work_centers[i]
                order = orders[i]

                if wc not in temp_dict:
                    temp_dict[wc] = set()
                temp_dict[wc].add(order)

            # 计算每个 Work Center 的唯一 Order 数量
            for wc, unique_orders in temp_dict.items():
                work_center_orders_count[wc] = len(unique_orders)
    return work_center_orders_count
    # for key,value in work_center_orders_count.items():
    #     print(key, value)
def read_excel_to_dict_openpyxl(file_path, sheet_name=None):
    """
    使用openpyxl读取Excel表格并返回字典格式
    """
    wb = load_workbook(filename=file_path, read_only=True)

    if sheet_name is None:
        ws = wb.active
    else:
        ws = wb[sheet_name]

    result = {}
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))

    for col_idx, col_name in enumerate(header_row):
        if col_name is None or str(col_name).strip() == '':
            continue

        col_key = str(col_name).strip()
        col_values = []
        for row in data_rows:
            if col_idx < len(row):
                value = row[col_idx]
                if value is not None:
                    col_values.append(value)
        result[col_key] = col_values

    wb.close()
    return result


def get_priority_batch(file_path, sheet_name=None):
    pass
    res_work_sheet = read_excel_to_dict_openpyxl(file_path, sheet_name=sheet_name)

    '''
    count work center order qty
    
    '''
    work_center_order_qty = count_work_center_order(res_work_sheet)

    '''
    count work center trolley number
    '''

    priority = res_work_sheet['Work center']

    priority = remove_duplicate(priority)


    print('find work center with sequence:'+str(len(priority)))

    res_trolley = read_excel_to_dict_openpyxl('../未来流量 模拟 setting(D+S).xlsx',sheet_name='trolley')

    # trolley capacity key = work center value = capacity
    trolley_capacity = defaultdict(str)

    #priority_trolley_dict key = priority(sorted),value = trolley_capacity
    priority_trolley_dict = defaultdict(str)

    for i in range(len(res_trolley['Work center'])):
        key = res_trolley['Work center'][i]
        trolley_capacity[key] = str(res_trolley['capacity'][i])
    print('-'*40)
    print(len(trolley_capacity.keys()))

    for work_center in priority:
        # print(work_center)
        # print(trolley_capacity[work_center])
        flag = 0
        for key in trolley_capacity.keys():
            if work_center in key:
                priority_trolley_dict[work_center] = trolley_capacity[key]
                flag = 1
        if flag == 0:
            print(f'cannot find {work_center} in trolley table, capacity 20 by default')
            priority_trolley_dict[work_center] = '20'
    #
    '''
    priority_trolley_dict key = work center value = trolley qty
    '''
    # for k,v in priority_trolley_dict.items():
    #     print(k,v)



    '''
    count batch for each work center
    '''

    work_center_batch_qty = defaultdict(str)

    for i in priority:
        work_center_batch_qty[i] = str(math.ceil(work_center_order_qty[i]/int(priority_trolley_dict[i])))




    return work_center_batch_qty,priority
if __name__ == '__main__':
    get_priority_batch('../未来流量 模拟 setting(D+S).xlsx','D52+S49')